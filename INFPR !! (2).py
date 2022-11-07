import openpyxl
from openpyxl.styles import Alignment
import pyautogui
import time
from openpyxl.styles import Border, Side
import re
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl import styles
def entry(strtrw,qty,remain):
    f1=open("inspection_qp.txt","r")
    saving=f1.read()
    
    workbook = openpyxl.load_workbook(saving)
    worksheet = workbook["QP"]
    worksheet2 = workbook["INFPR 11-25"]
    workbook6= openpyxl.load_workbook(r"C:\\Solution\INFPR\requirements.xlsx")
    worksheet6 = workbook6.active
    slno=[]
    specifications=[]
    drw_tol=[]
    tolplus=[]
    tolminus=[]
    tol=[]
    tolfinal=[]
    parameters=[]
    skip=[]
    required=[]
    qty_1=qty

    def check(start_rows,required,a):
        b = []
        for i in range(start_rows,start_rows+required):
            b.append(i)
        correction = []
        for i in b:
            if i not in a:
                correction.append(1)
            else:
                correction.append(0)
        if 0 in correction:
            return 0
        else:
            return 1
    def increase(start_rows,skip_1):
        try:
            list = skip_1
            a = start_rows
            start_rows = list[min(range(len(list)),key=lambda i: abs(list[i]-start_rows))]
            if start_rows == 30:
                start_rows = start_rows + 8
                return start_rows
            elif a<=start_rows:
                return start_rows+3
            else:
                start_rows = list[list.index(start_rows)+1]
                start_rows = start_rows+3
                return start_rows
        except:
            return 0
    def final(start_rows,skip_1):
        try:
            list = skip_1
            start_rows = list[min(range(len(list)),key=lambda i: abs(list[i]-start_rows))]
            start_rows = list[list.index(start_rows)+1]
            return start_rows
        except ValueError:
            return 0
    def check_pro(start_rows,newlist_1,skip_1,specifications,required,j):
        try:
        
            if(specifications[j+1] is not None and newlist_1[j+1] is None):
                require = [required[j]]
                for x in range(1,len(newlist_1)):
                    try:
                        if newlist_1[j+x] is not None:
                            break
                        else:
                            require.append(required[j+x])
                    except IndexError:
                        break
                
                if check(start_rows,sum(require),skip_1) == 0:
                    list=[x,1]
                    return list
                else:
                    list = [1,0]
                    return list
            elif newlist_1 == specifications:
                return check_pro(start_rows,slno,skip_1,specifications,required,j)
            else:
                list=[1,0]
                return list
        except IndexError:
            print("Final Value")
            list = [0,0]
            return list
        
    for i in range(30, worksheet.max_row, 29):
        skip.append(i)
        skip.append(i + 1)
        skip.append(i + 2)
        skip.append(i + 3)
    k = 1
    column=1
    row=2
    newlist=[slno,parameters,specifications,drw_tol,tolplus,tolminus,required]
    points = [1, 2, 3, 6, 14, 15, 16]
    m=0
    for j in points:
        for i in range(9, worksheet.max_row):
            if worksheet.cell(i, 2).value == "NOTES":
                break
            
            if i not in skip and worksheet.cell(i,16).value!="!!!":
                if j==16:
                    
                    newlist[m].append(int(worksheet.cell(i, j).value))
                else:
                    
                    newlist[m].append(str(worksheet.cell(i, j).value))

        m = m + 1
    for i in newlist:
        try:
            while True:
                i.remove("!!!")
        except ValueError:
            pass
    
    if strtrw==4:
        skip_1=[38,39,40]
        for i in range(73,worksheet2.max_row,35):
            skip_1.append(i)
            skip_1.append(i+1)
            skip_1.append(i+2)   
    else:
        skip_1=[strtrw+32,strtrw+33,strtrw+34]
        for i in range(strtrw+67,worksheet2.max_row,35):##############
            skip_1.append(i)
            skip_1.append(i+1)
            skip_1.append(i+2)
    tol=[]
    for i in range (0,len(tolplus)):
        if (tolplus[i] is None or tolplus[i] == "None") and (tolminus[i] is None or tolminus[i] == "None"):
            tol.append(None)
        else:
            if tolplus[i] is None or tolplus[i] == "None":
                tolplus[i] = "-"
            if tolminus[i] is None or tolminus[i] == "None":
                tolminus[i] = "-"
            
            if tolplus[i] == "-" and tolminus[i] != "-":
                tol.append(str(tolplus[i])+"/"+str(tolminus[i]))
            if tolminus[i] == "-" and tolplus[i] != "-":
                tol.append(str(tolplus[i])+"/"+str(tolminus[i]))
            if tolplus[i] == "-" and tolminus[i] == "-":
                tol.append("-")
            if tolplus[i] != "-" and tolminus[i] != "-":
                tol.append(str(tolplus[i])+"/"+str(tolminus[i]))

    for i in range(0, len(drw_tol)):
        if tol[i] is "-":
            tolfinal.append(drw_tol[i])
        else:
            tolfinal.append(tol[i])
    tol_min_final=[]
    tol_max_final=[]
    for  i in range(0,len(tolfinal)):
        if tolfinal[i] != "-":
            result=re.search("(.*)/(.*)",tolfinal[i])
            if result.group(2)=="-":
                tol_min_final.append(0)
            if result.group(1)=="-":
                tol_max_final.append(0)
            if result.group(2)!="-":
                if "'" in result.group(2):
                    a=result.group(2)[0:-1]
                    tol_min_final.append(float(a))
                else:
                    tol_min_final.append(float(result.group(2)))
            if result.group(1)!="-":
                if "'" in result.group(1):
                    s=result.group(1)[0:-1]
                    tol_max_final.append(float(s))
                else:
                    tol_max_final.append(float(result.group(1)))
        else:
            tol_min_final.append("-")
            tol_max_final.append("-")
    c=["0","1","2","3","4","5","6","7","8","9","."]
    specifications_new=[]
    minimum_value=[]
    maximum_value=[]
    for i in range(0,len(specifications)):
       
        if tolfinal[i]!="-":
            r=""
            if parameters[i] =="Angle" or parameters[i]=="Chamfer Angle":
                if "°" in specifications[i]and "'" in specifications[i]:
                    res=re.search("(.*)°(.*)'",str(specifications[i]))
                    a=float(res.group(1))+(float(res.group(2))/60)
                    angle=float(a)+(tol_min_final[i]/60)
                    whole=int(angle)
                    decimel=round((angle-int(angle))*60)
                    minimum_value.append(str(whole)+"°"+str(decimel)+"'")
                    angle=float(a)+(tol_max_final[i]/60)
                    whole=int(angle)
                    decimel=round((angle-int(angle))*60)
                    maximum_value.append(str(whole)+"°"+str(decimel)+"'")
                elif  "°" in specifications[i] and "'"  not in specifications[i]:
                    res1=re.search("(.*)°",str(specifications[i]))
                    a=float(res1.group(1))
                    minimum_value.append(str(a)+"°"+str(tol_min_final[i])+"'")
                    maximum_value.append(str(a)+"°"+str(tol_max_final[i])+"'")
            else:
                for j in range(0,len(specifications[i])):
                    if specifications[i][j] in c:
                        r=r+specifications[i][j]
                try:
                    specifications_new.append(float(r))
                    minimum_value.append(round(float(r)+tol_min_final[i],3))
                    maximum_value.append(round(float(r)+tol_max_final[i],3))
                except:
                    minimum_value.append("-")
                    maximum_value.append("-")
                
        else:
            minimum_value.append("-")
            maximum_value.append("-")
        
    newlist_1 = [slno, specifications, drw_tol,minimum_value,maximum_value]
    position=[]
    for i in range(0,len(slno)):
        if slno[i] is not None:
            position .append(i)

            
            
    new_re=[]
    for j in range(0,len(position)):
        m=0
        if position[j]==position[-1]:
            for x in range(position[j],len(required)):
                m=m+required[x]
        else:
            
            for i in range(position[j],position[j+1]):
                m=m+required[i]
        new_re.append(m)



    new_req=[]
    j=0
    for i in range(0,len(slno)):
            if slno[i] is not None:
                    new_req.append(new_re[j])
                    j=j+1
            else:
                    new_req.append(0)
    col=[1,2,3,22,23]
    e=0
    for i in range(0,len(newlist_1)):
        start_rows = strtrw####
        j = 0
        m = 0
        while(j<len(newlist_1[i])):
            list = []
            list = check_pro(start_rows,newlist_1[i],skip_1,specifications,required,j)
            if list[1]:
                try:
                    final_start_row = increase(start_rows,skip_1)
                    final_points = final(start_rows,skip_1)
                    print(start_rows,final_points)
                    for z in range(start_rows,final_points):
                        if z not in skip_1:
                            worksheet2.cell(z,col[e]).value = "!"
                            print(z,"fdsfdgdfg")
                    if newlist_1[i] is specifications:
                        print(final_start_row,"check")
                        worksheet2.cell(final_start_row,col[e]).value = newlist_1[i][j]
                        worksheet2.cell(final_start_row,col[e]).alignment = Alignment(horizontal = "center",vertical="center")
                        final_start_row = final_start_row + required[j]
                        
                        for x in range(1,list[0]):
                            worksheet2.cell(final_start_row,col[e]).value = newlist_1[i][j+x]
                            worksheet2.cell(final_start_row,col[e]).alignment = Alignment(horizontal = "center",vertical="center")
                            final_start_row = final_start_row+required[j+x]
                               
                        start_rows = final_start_row
                    else:
                        worksheet2.cell(final_start_row,col[e]).value = newlist_1[i][j]
                        worksheet2.cell(final_start_row,col[e]).alignment = Alignment(horizontal = "center",vertical="center")
                        start_rows = final_start_row + required[j]+required[j+1]
                    j = j+list[0]
                    while(True):
                        if new_req[m+1]==0:
                            m+=1
                        else:
                            m+=1
                            break
                    
                except IndexError:
                    print("Final_Value")
            else:
                try:
                    if check(start_rows,required[j],skip_1):
                        
                        worksheet2.cell(start_rows,col[e]).value=newlist_1[i][j]
                        worksheet2.cell(start_rows,col[e]).alignment = Alignment(horizontal = "center",vertical="center")
                        start_rows=start_rows+required[j]
                        j+=1
                    else:
                        
                        if start_rows not in skip_1:
                            worksheet2.cell(start_rows,col[e]).value = "!"
                            worksheet2.cell(start_rows,col[e]).alignment = Alignment(horizontal = "center",vertical="center")
                        start_rows+=1
                    m += 1
                except IndexError:
                    break
        e=e+1
    last_value = start_rows
    skip_2=[38,39,40]
    for i in range(73,worksheet2.max_row,35):
        skip_2.append(i)
        skip_2.append(i+1)
        skip_2.append(i+2)   
  
    for i in range(4,last_value):
   
        if worksheet2.cell(i,22).value is None and i not in skip_2: 
            worksheet2.cell(i,22).value = worksheet2.cell(i-1,22).value
        if worksheet2.cell(i,23).value is None and i not in skip_2:
            worksheet2.cell(i,23).value = worksheet2.cell(i-1,23).value
    font1=Font(bold=True,underline='single')

    for i in range(4,last_value):
    
        a=""
        if worksheet2.cell(i,22).value!="-" and i not in skip_2:
        
            for j in range(4,19):
                a=chr(j+64)+str(i)
                b="V"+str(i)
                c="W"+str(i)
                worksheet2.conditional_formatting.add(a,CellIsRule(operator='notBetween', formula=[b,c], stopIfTrue=True, font=font1))
    def add(d,skip_1):
        max_value = d[-1]
        for i in skip_1:
            if i>= max_value:
                break
            if i not in d:
                d.append(i)
        d.sort()
        return d
    col=[1,2,3]
    for j in col:
        dummy=[]
        for i in range(strtrw,last_value):
            if worksheet2.cell(i,j).value is not None:
                dummy.append(i)
        dummy = add(dummy,skip_1)
        try:
            for x in range(0,len(dummy)):
                if dummy[x] == dummy[-1]:
                    worksheet2.merge_cells(start_row=dummy[x], start_column=j, end_row=last_value-1,
                                              end_column=j)
                else:
                    worksheet2.merge_cells(start_row=dummy[x], start_column=j, end_row=dummy[x+1]-1,
                                               end_column=j)
        except IndexError:
            print("Reached")

    for i in range(1,worksheet2.max_row):
        for j in range (1,worksheet2.max_column):
            if (worksheet2.cell(i,j).value == "!"):
                worksheet2.cell(i,j).value = None

    a=worksheet6.cell(5,3).value
    if a is None:
        a=""

    part=worksheet6.cell(5,2).value
    c=len(str(part))
    if remain!=0:
        if strtrw==4:
            y=37
        else:
            y=35
        quantum=qty_1
        j = strtrw
        while(j<last_value):
            
            if j==4:
                y=37
            else:
                y=35
            qty_1=quantum
            for i in range(4,4+remain):
                
                worksheet2.cell(j-2,i).value=worksheet6.cell(7, 2).value
                worksheet2.cell(j-1,i).value=str(qty_1).zfill(c)+a
                print(qty_1)
                qty_1= qty_1+1
            qty = qty_1
            j = j + y


    else:
        if strtrw==4:
            y=37
        else:
            y=35
        qty_1 = qty
        j = strtrw
        while(j<last_value):
            if j==4:
                y=37
            else:
                y=35
            for i in range(4,19):
                
                worksheet2.cell(j-2,i).value=worksheet6.cell(7, 2).value
                worksheet2.cell(j-1,i).value=str(qty).zfill(c)+a   
                qty= qty+1
            qty = qty_1
            j = j + y
    notes_value=last_value
    
    for i in range(0,len(skip_1)):
        if last_value-skip_1[i] < 0:
            last_value=skip_1[i]
            break
    thin = Side(style="thin")
    c=0
    for i in range(strtrw,last_value):
        if worksheet2.cell(i,2).value is not None and worksheet2.cell(i,2).value != "Specification  in MM":
            for j in range(4, 19):
                worksheet2.cell(i+required[c], j).border = Border(top=thin,left = thin, right = thin)
                worksheet2.cell(i-1+required[c], j).border = Border( left = thin, right = thin)
            c+=1
    workbook.save(saving)
    workbook.close()
    
    return (last_value+3,qty,notes_value)

start_row = 4
f1=open("inspection_qp.txt","r")
saving=f1.read()
workbook5= openpyxl.load_workbook(r"C:\\Solution\INFPR\requirements.xlsx")
worksheet5 = workbook5.active
qty=int(worksheet5.cell(3, 2).value)-10
if qty>15:
    for i in range(11,qty,15):#################
        if(i+4)>qty:
            break
        print(start_row,i)
        (start_row,work,notes_value)=entry(start_row,i,0)
        
        
    if (qty)%15!=0:
        work=(((qty)//15)*15)+1+10
        print(start_row,work,(qty)%15)
        (start_row,work,notes_value)=entry(start_row,work,(qty)%15)
else:
    (start_row,work,notes_value)=entry(start_row,11,((qty)%15))
    print(start_row,work,notes_value)
workbook = openpyxl.load_workbook(saving)
worksheet = workbook["QP"]
worksheet2 = workbook["INFPR 11-25"]

skip=[]
skip_1=[start_row-1,start_row-2,start_row-3]
for i in range(30, worksheet.max_row, 29):
    skip.append(i)
    skip.append(i + 1)
    skip.append(i + 2)
    skip.append(i + 3)
for i in range(9, worksheet.max_row):
        if worksheet.cell(i, 2).value == "NOTES":
            start_notes=i
            break
nnotes=[]
ncomments=[]
nslno=[]
nspecs = []
ntol =[]
ndrg = []
nleast = []
last_value_2=notes_value
for i in range(start_notes,worksheet.max_row):
    if i in skip:
        pass
    elif worksheet.cell(i,2).value is None:
        break
    else:
        nnotes.append(worksheet.cell(i,2).value)
        nslno.append(worksheet.cell(i,1).value)
        ncomments.append(worksheet.cell(i,8).value)
        nspecs.append(worksheet.cell(i,3).value)
        ntol.append(worksheet.cell(i,6).value)
        ndrg.append(worksheet.cell(i,7).value)
        nleast.append(worksheet.cell(i,12).value)
for x in skip_1:
    if i<x:
        worksheet2.cell(x-2,24).value="$"
        break
j=0
border_start = last_value_2
thin = Side(style="thin")
for i in range(1,worksheet2.max_row):
    if last_value_2+i in skip_1:
        pass
    elif j == len(nslno):
        break
            
    else:
        thin = Side(style="thin")
        worksheet2.cell(last_value_2+i,1).value=nslno[j]
        worksheet2.cell(last_value_2+i,2).value=nnotes[j]
        worksheet2.cell(last_value_2+i,3).value=nspecs[j]
        worksheet2.cell(last_value_2+i,4).value=ntol[j]
        worksheet2.cell(last_value_2+i,5).value=ncomments[j]
        worksheet2.merge_cells(start_row=last_value_2+i,start_column=5,end_row=last_value_2+i,end_column=8)
        worksheet2.cell(last_value_2+i,9).value=nleast[j]
        row_height=(len(nnotes[j])//40)+1
        print(row_height)
        worksheet2.row_dimensions[last_value_2+i].height=row_height*18
        if nspecs[j] is None and ntol[j] is None and ncomments[j] is None and nleast[j] is None:
            worksheet2.unmerge_cells(start_row=last_value_2+i,start_column=5,end_row=last_value_2+i,end_column=8)
            worksheet2.merge_cells(start_row=last_value_2+i,start_column=2,end_row=last_value_2+i,end_column=9)
        elif nspecs[j] is None:
            worksheet2.merge_cells(start_row=last_value_2+i,start_column=2,end_row=last_value_2+i,end_column=3)
        j+=1
for i in range (border_start,last_value_2+1+i):
    for j in range(1,9):
        worksheet2.cell(i, j).border = Border(top=thin,left = thin, right = thin,bottom=thin)
                

workbook.save(saving)
pyautogui.alert("INFPR","INFPR IS READY")
