import os
from os import path
import os.path
import shutil
from win32com.client import Dispatch
import time
import openpyxl

path=r"C:\Solution"
if os.path.isdir(path)==False:
    os.makedirs(path,exist_ok=True)
source = "Print_Template.xlsx"
destination = r"C:\Solution"
shutil.copy(source, destination)   
path2 = "C:\Solution\Print_Template.xlsx"
xl = Dispatch("Excel.Application")
xl.Visible = False
f1 = open("print_sheets.txt","r")
a=f1.readlines()
file_name=a[len(a)-1]
a.remove(file_name)
file_name = file_name.strip("\n")
img = openpyxl.drawing.image.Image('qualityplan.png')

for i in a:
    wb1 = xl.Workbooks.Open(Filename=file_name)
    wb2 = xl.Workbooks.Open(Filename=path2)

    ws1 = wb1.Worksheets[i.strip("\n")]
    ws1.Copy(Before=wb2.Worksheets(1))

    wb2.Close(SaveChanges=True)
    wb1.Close(SaveChanges=True)

    workbook = openpyxl.load_workbook("C:\Solution\Print_Template.xlsx")
    worksheet = workbook[i.strip("\n")]
    
    if i == "QP\n":
        img.anchor = 'B1'
        worksheet.add_image(img)
        column = 16
        final_width = 77
    elif i== "NCR\n":
        img.anchor = 'A1'
        worksheet.add_image(img)
        column = 24
        final_width = 86
    elif i=="INFPR\n" or i=="REPORT\n":
        img.anchor = 'B1'
        worksheet.add_image(img)
        column = 24
        final_width = 84
    else:
        column = 24
        final_width = 84
    for j in range(1,worksheet.max_row):
        if worksheet.cell(j,column).value == "$":
            break
    area = "A1:"+chr(final_width)+str(j)
    worksheet.print_area=area
    print(area)
    workbook.save(path2)
    

std=workbook['Sheet1']
workbook.remove_sheet(std)
workbook.save(path2)

