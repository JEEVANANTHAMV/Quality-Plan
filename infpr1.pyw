import openpyxl
from openpyxl.styles import Alignment
from subprocess import call
import subprocess
from openpyxl.styles import Border, Side
import re

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl import styles
f1=open("inspection_qp.txt","r")
saving=f1.read()
print(saving)
f1.close()
workbook = openpyxl.load_workbook(saving)

worksheet = workbook["QP"]
worksheet2 = workbook["INFPR"]
worksheet3 = workbook["INFPR 11-25"]
workbook1 = openpyxl.load_workbook(r"C:\\Solution\INFPR\requirements.xlsx")
worksheet1 = workbook1.active

worksheet2.cell(3, 1).value = "PROJECT : " + worksheet.cell(4, 3).value

worksheet2.cell(3, 4).value = "PART NAME : " + str(worksheet.cell(5, 3).value)
worksheet2.cell(3, 13).value = worksheet.cell(6, 3).value
worksheet2.cell(4, 4).value = "ASSY NAME : " + str(worksheet.cell(7, 3).value)
worksheet2.cell(4, 9).value = "MATL. SPEC : " + str(worksheet.cell(6, 10).value)
worksheet2.cell(4, 1).value = "Po No :" + str(worksheet1.cell(2, 2).value)
worksheet2.cell(5, 1).value = "Offered Qty :" + str(worksheet1.cell(3, 2).value) + "No's"
asd=worksheet1.cell(5, 3).value
if asd is None:
    asd=""
worksheet2.cell(5, 4).value = "PART ID NO's "+worksheet1.cell(7, 2).value + str(worksheet1.cell(5, 2).value)+asd + "To" + worksheet1.cell(7,
                                                                                                                 2).value + str(worksheet1.cell(
    6, 2).value)+asd
worksheet2.cell(2, 19).value = "Wo No:" + str(worksheet1.cell(4, 2).value)
worksheet2.cell(6,19).value="Inst.ID:"
worksheet2.cell(36,19).value="Inst.ID:"
worksheet2.cell(6,20).value="Remarks"
worksheet2.cell(36,20).value="Remarks"
exception=[]
str1 = worksheet1.cell(8,2).value#############
try:
    exception = list(str1.split(","))
    ex=[]
    for i in exception:
        ex.append(int(i))
    exception=[]
    exception = ex
except:
    print("")
totalqty=int(worksheet1.cell(3,2).value)-len(exception)
if int(totalqty) <=10:
    qty=int(totalqty)
else:
    qty=10

slno=[]
specifications=[]
parameters = []
drw_tol=[]
tolplus=[]
tolminus=[]
tol=[]
tolfinal=[]
instrument=[]
leastcount=[]
skip=[]
required=[]

def check(start_rows,required,a):
    b = []
    for i in range(start_rows,start_rows+required):
        b.append(i)
    correction = []
    for i in b:
        if i not in a:
            correction.append(1)
        else:
            correction.append(0)
    if 0 in correction:
        return 0
    else:
        return 1
def increase(start_rows,skip_1):
    try:
        list = skip_1
        a = start_rows
        start_rows = list[min(range(len(list)),key=lambda i: abs(list[i]-start_rows))]
        if start_rows == 30:
            start_rows = start_rows + 8
            return start_rows
        elif a<=start_rows:
            return start_rows+3
        else:
            start_rows = list[list.index(start_rows)+1]
            start_rows = start_rows+3
            return start_rows
    except:
        return 0
def final(start_rows,skip_1):
    try:
        list = skip_1
        start_rows = list[min(range(len(list)),key=lambda i: abs(list[i]-start_rows))]
        if start_rows == 30:
            return start_rows
        else:
            start_rows = list[list.index(start_rows)+1]
            return start_rows
    except:
        return 0
def check_pro(start_rows,newlist_1,skip_1,specifications,required,j):
    try:
    
        if(specifications[j+1] is not None and newlist_1[j+1] is None):
            require = [required[j]]
            
            for x in range(1,len(newlist_1)):
                try:
                    if newlist_1[j+x] is not None:
                        break
                    else:
                        require.append(required[j+x])
                except IndexError:
                    break
            if check(start_rows,sum(require),skip_1) == 0:
                list=[x,1]
                return list
            else:
                list = [1,0]
                return list
        elif newlist_1 == specifications:
            return check_pro(start_rows,slno,skip_1,specifications,required,j)
        else:
            list=[1,0]
            return list
    except IndexError:
        print("Final Value")
        list = [0,0]
        return list
    
for i in range(30, worksheet.max_row, 29):
    skip.append(i)
    skip.append(i + 1)
    skip.append(i + 2)
    skip.append(i + 3)
k = 1
column=1
row=2

newlist=[slno,parameters,specifications,drw_tol,instrument,leastcount,tolplus,tolminus,required]
    
points = [1,2,  3, 6,  8, 12, 14, 15,16]
m=0
for j in points:
    for i in range(9, worksheet.max_row):
        if worksheet.cell(i, 2).value == "NOTES":
            start_notes=i
            break
        if i not in skip and worksheet.cell(i,16).value!= "!!!":
            if m==8:
                newlist[m].append(int(worksheet.cell(i, j).value))
            else:   
                newlist[m].append(str(worksheet.cell(i, j).value))

    m = m + 1
for i in newlist:
    try:
        while True:
            i.remove("!!!")
    except ValueError:
        pass
tol_min=[]
tol_max=[]
nnotes=[]
ncomments=[]
nslno=[]
nspecs = []
ntol =[]
ndrg = []
nleast = []
for i in range(start_notes,worksheet.max_row):
    if i in skip:
        pass
    elif worksheet.cell(i,2).value is None:
        break
    else:
        nnotes.append(worksheet.cell(i,2).value)
        nslno.append(worksheet.cell(i,1).value)
        ncomments.append(worksheet.cell(i,8).value)
        nspecs.append(worksheet.cell(i,3).value)
        ntol.append(worksheet.cell(i,6).value)
        ndrg.append(worksheet.cell(i,7).value)
        nleast.append(worksheet.cell(i,12).value)
skip_1=[30,31,32,33,34,35,36,37,72,73,74]
for i in range(107,worksheet2.max_row,35):
    skip_1.append(i)
    skip_1.append(i+1)
    skip_1.append(i+2)
tol=[]
for i in range (0,len(tolplus)):
    if (tolplus[i] is None or tolplus[i] == "None") and (tolminus[i] is None or tolminus[i] == "None"):
        tol.append(None)
    else:
        if tolplus[i] is None or tolplus[i] == "None":
            tolplus[i] = "-"
        if tolminus[i] is None or tolminus[i] == "None":
            tolminus[i] = "-"
        
        if tolplus[i] == "-" and tolminus[i] != "-":
            tol.append(str(tolplus[i])+"/"+str(tolminus[i]))
        if tolminus[i] == "-" and tolplus[i] != "-":
            tol.append(str(tolplus[i])+"/"+str(tolminus[i]))
        if tolplus[i] == "-" and tolminus[i] == "-":
            tol.append("-")
        if tolplus[i] != "-" and tolminus[i] != "-":
            tol.append(str(tolplus[i])+"/"+str(tolminus[i]))

for i in range(0, len(drw_tol)):
    if tol[i] is "-":
        tolfinal.append(drw_tol[i])
    else:
        tolfinal.append(tol[i])
tol_min_final=[]
tol_max_final=[]
for  i in range(0,len(tolfinal)):
    if tolfinal[i] != "-":
        result=re.search("(.*)/(.*)",tolfinal[i])
        if result.group(2)=="-":
            tol_min_final.append(0)
        if result.group(1)=="-":
            tol_max_final.append(0)
        if result.group(2)!="-":
            if "'" in result.group(2):
                f=result.group(2)[0:-1]
                tol_min_final.append(float(f))
            else:
                tol_min_final.append(float(result.group(2)))
        if result.group(1)!="-":
            if "'" in result.group(1):
                s=result.group(1)[0:-1]
                tol_max_final.append(float(s))
            else:
                tol_max_final.append(float(result.group(1)))
    else:
        tol_min_final.append("-")
        tol_max_final.append("-")
c=["0","1","2","3","4","5","6","7","8","9","."]
specifications_new=[]
minimum_value=[]
maximum_value=[]
for i in range(0,len(specifications)):
   
    if tolfinal[i]!="-" :
        r=""
        if parameters[i] =="Angle" or parameters[i]=="Chamfer Angle":
            if "°" in specifications[i]and "'" in specifications[i] :
                res=re.search("(.*)°(.*)'",str(specifications[i]))
                a=float(res.group(1))+(float(res.group(2))/60)        
                angle=float(a)+(tol_min_final[i]/60)
                whole=int(angle)
                decimel=round((angle-int(angle))*60)
                minimum_value.append(str(whole)+"°"+str(decimel)+"'")
                angle=float(a)+(tol_max_final[i]/60)
                whole=int(angle)
                decimel=round((angle-int(angle))*60)
                maximum_value.append(str(whole)+"°"+str(decimel)+"'")
            elif  "°" in specifications[i] and "'"  not in specifications[i]:
                res1=re.search("(.*)°",str(specifications[i]))
                a=float(res1.group(1))
                minimum_value.append(str(a)+"°"+str(tol_min_final[i])+"'")
                maximum_value.append(str(a)+"°"+str(tol_max_final[i])+"'")
        else:
                for j in range(0,len(str(specifications[i]))):
                    if (specifications[i][j]) in c:
                        r=r+(specifications[i][j])
                try:
                    specifications_new.append(float(r))  
                    minimum_value.append(round(float(r)+tol_min_final[i],3))
                    maximum_value.append(round(float(r)+tol_max_final[i],3))
                except:
                    minimum_value.append("-")
                    maximum_value.append("-")
    else:
        minimum_value.append("-")
        maximum_value.append("-")
        
newlist_1 = [slno, specifications, drw_tol, instrument, leastcount,minimum_value,maximum_value]
col=[1,2,3,4,8,22,23]
e=0
start_rows=8
position=[]
for i in range(0,len(slno)):
    if slno[i] is not None:
        position .append(i)

        
        
new_re=[]
for j in range(0,len(position)):
    m=0
    if position[j]==position[-1]:
        for x in range(position[j],len(required)):
            m=m+required[x]
    else:
        
        for i in range(position[j],position[j+1]):
            m=m+required[i]
    new_re.append(m)



new_req=[]
j=0
for i in range(0,len(slno)):
	if slno[i] is not None:
		new_req.append(new_re[j])
		j=j+1
	else:
		new_req.append(0)
		
for i in range(0,len(newlist_1)):
    start_rows = 8
    j = 0
    m=0
    
    while(j<len(newlist_1[i])):
        
        list = []
        list = check_pro(start_rows,newlist_1[i],skip_1,specifications,required,j)
        if list[1]:
            try:
                final_start_row = increase(start_rows,skip_1)
                final_points = final(start_rows,skip_1)
                for z in range(start_rows,final_points):
                    if z not in skip_1:
                        worksheet2.cell(z,col[e]).value = "!"
                if newlist_1[i] is specifications:
                    worksheet2.cell(final_start_row,col[e]).value = newlist_1[i][j]
                    worksheet2.cell(final_start_row,col[e]).alignment = Alignment(horizontal = "center",vertical="center")
                    final_start_row = final_start_row + required[j]
                    for x in range(1,list[0]):
                        worksheet2.cell(final_start_row,col[e]).value = newlist_1[i][j+x]
                        worksheet2.cell(final_start_row,col[e]).alignment = Alignment(horizontal = "center",vertical="center")
                        final_start_row = final_start_row+required[j+x]
                    start_rows = final_start_row
                    
                else:
                    worksheet2.cell(final_start_row,col[e]).value = newlist_1[i][j]
                    worksheet2.cell(final_start_row,col[e]).alignment = Alignment(horizontal = "center",vertical="center")
                    start_rows = final_start_row +new_req[m]
                j = j+list[0]
                while(True):
                    if new_req[m+1]==0:
                        m+=1
                    else:
                        m+=1
                        break
                
            except:
                print("Final_Value")
        else:
            try:
                if check(start_rows,required[j],skip_1):
                    worksheet2.cell(start_rows,col[e]).value=newlist_1[i][j]
                    worksheet2.cell(start_rows,col[e]).alignment = Alignment(horizontal = "center",vertical="center")
                    start_rows=start_rows+required[j]
                    j+=1
                    m = m + 1
                else:
                    if start_rows not in skip_1:
                        worksheet2.cell(start_rows,col[e]).value = "!"
                        worksheet2.cell(start_rows,col[e]).alignment = Alignment(horizontal = "center",vertical="center")
                    start_rows+=1
            except IndexError:
                break
    e=e+1

last_value = start_rows

for i in range(8,last_value):
   
    if worksheet2.cell(i,22).value is None and i not in skip_1:
        
        worksheet2.cell(i,22).value = worksheet2.cell(i-1,22).value
    if worksheet2.cell(i,23).value is None and i not in skip_1:
        worksheet2.cell(i,23).value = worksheet2.cell(i-1,23).value
font1=Font(bold=True,underline='single')

for i in range(8,last_value):
    
    a=""
    if worksheet2.cell(i,22).value!="-" and i not in skip_1:
        
        for j in range(9,19):
            a=chr(j+64)+str(i)
            b="V"+str(i)
            c="W"+str(i)
            worksheet2.conditional_formatting.add(a,CellIsRule(operator='notBetween', formula=[b,c], stopIfTrue=True, font=font1))
 
print("25")
def add(d,skip_1):
    max_value = d[-1]
    for i in skip_1:
        if i>= max_value:
            break
        if i not in d:
            d.append(i)
    d.sort()
    return d
col = [1,2,3,4,8]
for j in col:
    dummy=[]
    for i in range(8,last_value):
        if worksheet2.cell(i,j).value is not None:
            dummy.append(i)
    dummy = add(dummy,skip_1)
    for x in range(0,len(dummy)):
        if dummy[x] == dummy[-1]:
            if j!=4:
                worksheet2.merge_cells(start_row=dummy[x], start_column=j, end_row=last_value-1,
                                      end_column=j)
            else:
                worksheet2.merge_cells(start_row=dummy[x], start_column=j, end_row=last_value-1,
                                      end_column=j+3)
        else:
            if j != 4:
                worksheet2.merge_cells(start_row=dummy[x], start_column=j, end_row=dummy[x+1]-1,
                                       end_column=j)
            else: 
                worksheet2.merge_cells(start_row=dummy[x], start_column=j, end_row=dummy[x+1]-1,
                                       end_column=j+3)
for i in range(1,worksheet2.max_row):
    for j in range (1,worksheet2.max_column):
        if (worksheet2.cell(i,j).value == "!"):
            worksheet2.cell(i,j).value = None

a=worksheet1.cell(5,3).value
if a is None:
    a=""

part=worksheet1.cell(5,2).value
c=len(str(part))
#edited


if last_value < 36:
    
    i=9
    j=9
    while(1):
        if i==(9+qty):
            break
        elif i-8 in exception:
            i=i+1
            qty = qty+1
        else:
            worksheet2.cell(6,j).value=worksheet1.cell(7, 2).value
            worksheet2.cell(7,j).value=str(i-8).zfill(c)+a
            worksheet2.cell(7,j).alignment = Alignment(horizontal = "center",vertical="center")
            worksheet2.cell(6,j).alignment = Alignment(horizontal = "center",vertical="center")
            i = i+1
            j=j+1
            print(i,j)
            
if last_value<72 and last_value>36:
    i=9
    j=9
    
    while(1):
        if i==(9+qty):
            break
        elif i-8 in exception:
            i=i+1
            qty = qty+1
        else:
            worksheet2.cell(36,j).value=worksheet1.cell(7, 2).value
            worksheet2.cell(37,j).value=str(i-8).zfill(c)+a
            worksheet2.cell(37,j).alignment = Alignment(horizontal = "center",vertical="center")
            worksheet2.cell(36,j).alignment = Alignment(horizontal = "center",vertical="center")
            worksheet2.cell(6,j).value=worksheet1.cell(7, 2).value
            worksheet2.cell(7,j).value=str(i-8).zfill(c)+a
            worksheet2.cell(7,j).alignment = Alignment(horizontal = "center",vertical="center")
            worksheet2.cell(6,j).alignment = Alignment(horizontal = "center",vertical="center")
            j=j+1
if last_value>72:
    i=9
    j=9
    
    while(1):
        if i==(9+qty):
            break
        elif i-8 in exception:
            i=i+1
            qty = qty+1
        else:
            worksheet2.cell(36,j).value=worksheet1.cell(7, 2).value
            worksheet2.cell(37,j).value=str(i-8).zfill(c)+a
            worksheet2.cell(6,j).value=worksheet1.cell(7, 2).value
            worksheet2.cell(7,j).value=str(i-8).zfill(c)+a
            worksheet2.cell(37,j).alignment = Alignment(horizontal = "center",vertical="center")
            worksheet2.cell(36,j).alignment = Alignment(horizontal = "center",vertical="center")
            worksheet2.cell(7,j).alignment = Alignment(horizontal = "center",vertical="center")
            worksheet2.cell(6,j).alignment = Alignment(horizontal = "center",vertical="center")
            j=j+1
    for j in range(73,last_value,35):
        i=9
        k=9
        while(1):
            if i==(9+qty):
                break
            elif i-8 in exception:
                i=i+1
                qty = qty+1
            else:
                worksheet2.cell(j,k).value=worksheet1.cell(7, 2).value
                worksheet2.cell(j,k).alignment = Alignment(horizontal = "center",vertical="center")
                worksheet2.cell(j+1,k).value=str(i-8).zfill(c)+a #####need to add the any alphabet
                worksheet2.cell(j+1,k).alignment = Alignment(horizontal = "center",vertical="center")
                k=k+1
#edited
thin = Side(style="thin")
c = 0
for i in range(8,last_value):
    if worksheet2.cell(i,2).value is not None and worksheet2.cell(i,2).value != "Specification  in MM":
        for j in range(9, 19):
            worksheet2.cell(i+required[c], j).border = Border(top=thin,left = thin, right = thin)
            worksheet2.cell(i-1+required[c], j).border = Border( left = thin, right = thin)
        c+=1
        
     

last_value_2=last_value
for i in range(30,38):
    try:
        skip_1.remove(i)
    except:
        print("Index")
        
last_value_1 = skip_1[min(range(len(skip_1)),key=lambda i: abs(skip_1[i]-last_value))]
if last_value_1 >last_value:
    last_value = last_value_1
else:
    last_value = skip_1[skip_1.index(last_value_1)+1]
    
for i in range(0,len(skip_1)):
    if(skip_1[i]<=last_value):
        a = skip_1[i]
        print(a)
        if (a in skip_1 and a+1 in skip_1 and a+2 in skip_1):
            continue

worksheet2.cell(a,24).value="$"
if int(worksheet1.cell(3,2).value)>10:
    ar=last_value/35
    if round(ar)>ar:
        ar=round(ar)
    else:
        ar=round(ar)+1
    aw=(int(worksheet1.cell(3, 2).value)-10)/(15)
    if round(int(worksheet1.cell(3, 2).value)-10)/(15)>aw:
        aw=round(aw)
    else:
        aw=round(aw)+1
        
    quantity=(aw*(ar))

    ws=workbook["INFPR 11-25"]
    a_new=424
    print("50")
    for x in range(0,quantity):
        
        for i in range(39,74):
            for j in range(1,21):
                if i ==40 and j==1:
                    ws.merge_cells(start_row=a_new-1,start_column=1,end_row=a_new,end_column=1)
                elif i ==40 and j==2:
                    ws.merge_cells(start_row=a_new-1,start_column=2,end_row=a_new,end_column=2)
                elif i ==40 and j==3:
                    ws.merge_cells(start_row=a_new-1,start_column=3,end_row=a_new,end_column=3)
                elif i ==40 and j==19:
                    ws.merge_cells(start_row=a_new-1,start_column=19,end_row=a_new,end_column=19)
                elif i ==40 and j==20:
                    ws.merge_cells(start_row=a_new-1,start_column=20,end_row=a_new,end_column=20)
                else:
                    ws.cell(a_new,j).value=ws.cell(i,j).value
                    ws.cell(a_new,j)._style = ws.cell(i,j)._style
                    
            a_new=a_new+1
        ws.merge_cells(start_row=a_new-1,start_column=1,end_row=a_new-1,end_column=4)
        
        ws.merge_cells(start_row=a_new-1,start_column=5,end_row=a_new-1,end_column=10)
        
        ws.merge_cells(start_row=a_new-1,start_column=11,end_row=a_new-1,end_column=15)
        
        ws.row_dimensions[a_new-1].height=30
    
    workbook.save(saving[:-5]+"-"+str(worksheet1.cell(7, 2).value)+str(worksheet1.cell(5, 2).value)+asd+".xlsx")
    saving=saving[:-5]+"-"+str(worksheet1.cell(7, 2).value)+str(worksheet1.cell(5, 2).value)+asd+".xlsx"
    f1=open("inspection_qp.txt","w")
    f1.write(saving)
    
    f1.close()
    cmd=["python","INFPR !!.py"]
    subprocess.Popen(cmd).wait()
else:
    j=0
    for i in range(1,worksheet2.max_row):
        if last_value_2+i in skip_1:
            pass
        elif j == len(nslno):
            break
            
        else:
            thin = Side(style="thin")
            worksheet2.cell(last_value_2+i,1).value=nslno[j]
            worksheet2.cell(last_value_2+i,2).value=nnotes[j]
            worksheet2.cell(last_value_2+i,3).value=nspecs[j]
            worksheet2.cell(last_value_2+i,4).value=ntol[j]
            worksheet2.cell(last_value_2+i,5).value=ncomments[j]
            worksheet2.merge_cells(start_row=last_value_2+i,start_column=5,end_row=last_value_2+i,end_column=8)
            worksheet2.cell(last_value_2+i,9).value=nleast[j]
            row_height=(len(nnotes[j])//40)+1
            print(row_height)
            worksheet2.row_dimensions[last_value_2+i].height=row_height*18
            if nspecs[j] is None and ntol[j] is None and ncomments[j] is None and nleast[j] is None:
                worksheet2.unmerge_cells(start_row=last_value_2+i,start_column=5,end_row=last_value_2+i,end_column=8)
                worksheet2.merge_cells(start_row=last_value_2+i,start_column=2,end_row=last_value_2+i,end_column=9)
            elif nspecs[j] is None:
                worksheet2.merge_cells(start_row=last_value_2+i,start_column=2,end_row=last_value_2+i,end_column=3)
            j+=1

    workbook.save(saving[:-5]+"-"+str(worksheet1.cell(7, 2).value)+str(worksheet1.cell(5, 2).value)+asd+".xlsx")
    saving=saving[:-5]+"-"+str(worksheet1.cell(7, 2).value)+str(worksheet1.cell(5, 2).value)+asd+".xlsx" 
    f1=open("inspection_qp.txt","w")
    f1.write(saving)
    
    f1.close()
              
            
        
