import openpyxl
import os
import pyautogui
from openpyxl.styles import Alignment
from datetime import datetime
today=datetime.now()
workbook = openpyxl.load_workbook(r"C:\\Solution\Editedqp\Temp1.xlsx")
sheet = workbook.active
input_1 = []
slnolist = []
for i in range(3, 17):
    cell = sheet.cell(i, 2)
    input_1.append(str(cell.value))
workbook = openpyxl.load_workbook(r"C:\\Solution\Editedqp\Temp2.xlsx")
sheet = workbook.active
parameters = []
specific_dim = []
instrument = []
drg_zone = []
drwtol = []
tolplus = []
tolminus = []
remarks = []
slno = []
comments = []
least_count = []
symbols = []
infpr=[]
nrows = sheet.max_row
for i in range(2, nrows + 1):
    slno.append(sheet.cell(i, 1).value)
    parameters.append(sheet.cell(i, 2).value)
    specific_dim.append(sheet.cell(i,3).value)
    drwtol.append(sheet.cell(i,4).value)
    drg_zone.append(sheet.cell(i, 5).value)
    instrument.append(sheet.cell(i, 6).value)
    least_count.append(sheet.cell(i, 7).value)
    remarks.append(sheet.cell(i,8).value)
    tolplus.append(sheet.cell(i, 9).value)
    tolminus.append(sheet.cell(i, 10).value)
    infpr.append(sheet.cell(i,11).value)
workbook = openpyxl.load_workbook(r"C:\\Solution\Editedqp\Temp3.xlsx")
sheet = workbook.active
max_value = sheet.max_row
nslno = []
nnotes = []
nspec=[]
ntol=[]
ndrg=[]

ncomments=[]
nleast=[]
for i in range(3, max_value + 1):
    nslno.append(sheet.cell(i, 1).value)
    nnotes.append(sheet.cell(i, 2).value)
    nspec.append(sheet.cell(i,3).value)
    ntol.append(sheet.cell(i,4).value)
    ndrg.append(sheet.cell(i,5).value)
    ncomments.append(sheet.cell(i,6).value)
    nleast.append(sheet.cell(i,7).value)
    
workbook1 = openpyxl.load_workbook("final_temp.xlsx")
worksheet1 = workbook1["QP"]
worksheet1.cell(4, 3).value = input_1[0]
worksheet1.cell(5, 3).value = input_1[1]
worksheet1.cell(6, 3).value = "DRG.No:" + input_1[2] + ",Rev:" + input_1[3] + ",ISSUE:" + input_1[4]
worksheet1.cell(7, 3).value = input_1[5]
worksheet1.cell(4, 10).value = input_1[6]
worksheet1.cell(5, 10).value = input_1[7]
worksheet1.cell(6, 10).value = input_1[8]
worksheet1.cell(7, 10).value = input_1[9]
##########Header
worksheet1.cell(30, 3).value = input_1[10]
worksheet1.cell(30, 8).value = input_1[11]
worksheet1.cell(30, 12).value = input_1[12]
worksheet1.cell(31, 3).value = input_1[13]
worksheet1.cell(32, 1).value = "PART NAME: " + input_1[1]
worksheet1.cell(32, 8).value = "DRG.No:" + input_1[2] + ",Rev:" + input_1[3] + ",ISSUE:" + input_1[4]
newlist = [slno, parameters, specific_dim, drwtol, drg_zone, instrument, least_count, remarks, tolplus, tolminus]
skip = []
for i in range(30, worksheet1.max_row, 29):
    skip.append(i)
    skip.append(i + 1)
    skip.append(i + 2)
    skip.append(i + 3)
def common_data(list1, list2): 
    result = 1
    for x in list1: 
        for y in list2: 
            if x == y:
                result = 0
                return result 
    return result 
def check(slno,j,skip,row):
    if row in skip:
        return 0 
    list1 = [row]
    row+=1
    for i in range(j+1,len(slno)):
        try:
            if slno[i] is None:
                list1.append(row)
                row+=1  
            else:
                break
        except:
            print("Final_Value")
    fine = common_data(list1,skip)
    return fine
i = 9
j = 0
max_size = len(slno)
k = [1,2,3,6,7,8,12,13,14,15,16]
while (j<max_size):
    print(j)
    allow = check(slno,j,skip,i)
    if allow == 0:
        if i not in skip:
            for z in range(0,len(k)):
                worksheet1.cell(i,k[z]).value= "!!!"
        i = i + 1
    else:
        worksheet1.cell(i, 1).value = slno[j]
        worksheet1.cell(i, 2).value = parameters[j]
        worksheet1.cell(i, 3).value = specific_dim[j]
        worksheet1.cell(i, 6).value = drwtol[j]
        worksheet1.cell(i, 7).value = drg_zone[j]
        worksheet1.cell(i, 8).value = instrument[j]
        worksheet1.cell(i, 12).value = least_count[j]
        worksheet1.cell(i, 13).value = remarks[j]
        worksheet1.cell(i, 14).value = tolplus[j]
        worksheet1.cell(i, 15).value = tolminus[j]
        worksheet1.cell(i, 16).value=infpr[j]
        i = i + 1
        j = j + 1
        if i - 1 in skip and i - 1 != 8:
            worksheet1.cell(i-4, 3).value = input_1[10]
            worksheet1.cell(i-4, 8).value = input_1[11]
            worksheet1.cell(i-4, 12).value = input_1[12]
            worksheet1.cell(i -3, 3).value = input_1[14]
            worksheet1.cell(i - 2, 1).value = "PART NAME: " + input_1[1]
            worksheet1.cell(i - 2, 8).value = "DRG.No:" + input_1[2] + ",Rev:" + input_1[3] + ",ISSUE:" + input_1[4]
            
final_value_row = i-1
p = 0
a = 0
if i in skip:
    i=i+4
worksheet1.unmerge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
worksheet1.unmerge_cells(start_row=i, start_column=8, end_row=i, end_column=11)
worksheet1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
worksheet1.merge_cells(start_row=i, start_column=8, end_row=i, end_column=13)
worksheet1.cell(i, 1).value = "Slno"
worksheet1.cell(i,1).alignment=Alignment(horizontal="left",vertical="center")
worksheet1.cell(i, 2).value = "NOTES"
worksheet1.cell(i,2).alignment=Alignment(horizontal="left",vertical="center")
worksheet1.cell(i, 8).value = "COMMENTS"
worksheet1.cell(i,8).alignment=Alignment(horizontal="left",vertical="center")
b=1
c=i+1

while (c<= len(nslno) + i + a):
    try:
        if c not in skip:
            worksheet1.cell(c, 1).value = nslno[p]
            worksheet1.cell(c,1).alignment=Alignment(horizontal="left",vertical="center")
            
            worksheet1.cell(c, 2).value = nnotes[p]
            worksheet1.cell(c, 3).value = nspec[p]
            worksheet1.cell(c, 6).value = ntol[p]
            worksheet1.cell(c, 7).value = ndrg[p]
            worksheet1.cell(c, 12).value = nleast[p]
            
            row_height=(len(nnotes[p])//40)+1
            worksheet1.row_dimensions[c].height=row_height*18
            worksheet1.cell(c,2).alignment=Alignment(horizontal="left",vertical="center")
            worksheet1.cell(c, 8).value = ncomments[p]
            worksheet1.cell(c,8).alignment=Alignment(horizontal="left",vertical="center")
            if nspec[p] is None and ntol[p] is None and ndrg[p] is None and ncomments[p] is None and nleast[p] is None:
                worksheet1.unmerge_cells(start_row = c, start_column = 3,end_row=c,end_column=5)
                worksheet1.unmerge_cells(start_row = c, start_column = 8,end_row=c,end_column=11)
                worksheet1.merge_cells(start_row = c, start_column = 2,end_row=c,end_column=13)
                worksheet1.cell(c, 1).value = nslno[p]
                worksheet1.cell(c,1).alignment=Alignment(horizontal="left",vertical="center")
                worksheet1.cell(c, 2).value = nnotes[p]
            elif nspec[p] is None or nspec[p]=="None":
                worksheet1.unmerge_cells(start_row = c, start_column = 3,end_row=c,end_column=5)
                worksheet1.merge_cells(start_row = c, start_column = 2 ,end_row=c,end_column=5)
            p = p + 1
        if c in skip:
            c=c+4
        elif c+1 in skip:
        
            c+=5
            a+= 5
        else:
            c+=1
    except IndexError:
        break
        
lastrow=c
for i in skip:
    if lastrow<i:
        ab=i
        break
worksheet1.cell(ab, 3).value = input_1[10]
worksheet1.cell(ab, 8).value = input_1[11]
worksheet1.cell(ab, 12).value = input_1[12]
worksheet1.cell(ab + 1, 3).value = input_1[13]
worksheet1.cell(ab + 2, 1).value = "PART NAME: " + input_1[1]
worksheet1.cell(ab+ 2, 8).value = "DRG.No:" + input_1[2] + ",Rev:" + input_1[3] + ",ISSUE:" + input_1[4]
worksheet1.cell(ab+ 1, 16).value ="$"
for j in k:
    i=9
    while(i<worksheet1.max_row):
        if (worksheet1.cell(i,j).value is None and i not in skip):
            strw = i-1
            endrw = i
            for m in range(i,worksheet1.max_row):
                if worksheet1.cell(m,j).value is not None or worksheet1.cell(m+1,2).value == "NOTES":
                    endrw = m
                    break
            if(worksheet1.cell(m+1,2).value == "NOTES"):
                break
            if j==3:
                for m in range(strw,endrw):
                   worksheet1.unmerge_cells(start_row = m, start_column = 3, end_row = m, end_column= 5)
                worksheet1.merge_cells(start_row=strw,start_column = 3,end_row = endrw-1,end_column= 5)
            elif j==8:
                for m in range(strw,endrw):
                   worksheet1.unmerge_cells(start_row = m, start_column = 8, end_row = m, end_column= 11)
                worksheet1.merge_cells(start_row=strw,start_column = 8,end_row = endrw-1,end_column= 11)
            else:
                worksheet1.merge_cells(start_row=strw,start_column = j,end_row = endrw-1,end_column= j)
            if(i==endrw):
                i = i +1
            else:
                i = endrw 
        elif (worksheet1.cell(i+1,2).value == "NOTES"):
            break
        else:
            i+=1
last_value = i
while(1):
    if last_value not in skip:
        break
    else:
        last_value-=1
p = last_value

for j in k:
    for i in range(last_value,0,-1):
        if(worksheet1.cell(i,j).value is not None):
            p = i
            break
    if p!= last_value:
        if j==3:
            for i in range(p,last_value+1):
                worksheet1.unmerge_cells(start_row = i, start_column = 3, end_row = i , end_column = 5)
            worksheet1.merge_cells(start_row = p, start_column = 3, end_row = last_value, end_column= 5 )
        elif j==8:
            for i in range(p,last_value+1):
                worksheet1.unmerge_cells(start_row = i, start_column = 8, end_row = i , end_column = 11)
            worksheet1.merge_cells(start_row = p, start_column = 8, end_row = last_value, end_column= 11 )
        else:
            worksheet1.merge_cells(start_row = p, start_column = j, end_row = last_value, end_column = j)
if worksheet1.cell(final_value_row,1).value is not None and worksheet1.cell(final_value_row-1,1).value is None:
    last_value = final_value_row - 1
    for j in k:
        for i in range(last_value,0,-1):
            if(worksheet1.cell(i,j).value is not None):
                p = i
                break
        if p!= last_value:
            if j==3:
                for i in range(p,last_value+1):
                    worksheet1.unmerge_cells(start_row = i, start_column = 3, end_row = i , end_column = 5)
                worksheet1.merge_cells(start_row = p, start_column = 3, end_row = last_value, end_column= 5 )
            elif j==8:
                for i in range(p,last_value+1):
                    worksheet1.unmerge_cells(start_row = i, start_column = 8, end_row = i , end_column = 11)
                worksheet1.merge_cells(start_row = p, start_column = 8, end_row = last_value, end_column= 11 )
            else:
                worksheet1.merge_cells(start_row = p, start_column = j, end_row = last_value, end_column = j)        
f3=open("inspection_qp.txt","r")
path=f3.read()
path_1=os.path.split(path)
for row in worksheet1.iter_rows():
    for cell in row:
        cell.alignment=Alignment(wrap_text=True,vertical='center')
workbook1.save(path_1[0]+"/"+str(today.strftime("%A %h %Y"))+input_1[1]+"-"+input_1[2]+"-"+input_1[3]+".xlsx")
pyautogui.alert("QP IS READY", " ")
